import sqlite3
import openpyxl

lists = sqlite3.connect('./Database/student.sqlite')
c = lists.cursor()
listinsheet = openpyxl.load_workbook(r'D:\grade\cet.xlsx')
datainlist = listinsheet.active  # 获取excel文件当前表格
data_truck = '''INSERT INTO cet(cetid,studentid,name,cet4_grade,cet6_grade) VALUES (?,?,?,?,?)'''
for row in datainlist.iter_rows(min_row=2, max_col=5, max_row=datainlist.max_row):
    # 使excel各行数据成为迭代器
    cargo = [cell.value for cell in row]  # 敲黑板！！使每行中单元格成为迭代器
    c.execute(data_truck, cargo)  # 敲黑板！写入一行数据到数据库中表mylist
lists.commit()
print(11)
lists.close()
